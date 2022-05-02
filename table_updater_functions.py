import string
from openpyxl.utils import *
from openpyxl import workbook, load_workbook
from table_updater_gui import UpdaterGui

def submit():
    
    workbook = load_workbook('workbook.xlsx')
    worksheet = workbook['sheet']
    table = worksheet.tables['table']

    #setting variables for the workbook, the worksheet and the table

    ref_1 = worksheet[table.ref][0][0]
    ref_2 = worksheet[table.ref][-1][-1]

    #setting variables for the two references cell of the table

    current_row = ref_2.row + 1
    current_column = ref_1.column

    #reference to start setting the data

    new_data = [app.entry_data1.get(), app.entry_data2.get(), app.entry_data3.get()]

    for i in range(3):
        worksheet.cell(row=current_row, column=current_column + i).value = new_data[i]

    #setting the new data
    
    current_ref = list(range_boundaries(table.ref))
    current_ref[-1] += 1
    table.ref = f"{get_column_letter(current_ref[0])}{current_ref[1]}:{get_column_letter(current_ref[2])}{current_ref[3]}"

    #setting the new table reference

    workbook.save('workbook.xlsx')

class FullUpdater(UpdaterGui):
    def add_command(self):
        self.button.configure(command=submit)

if __name__ == '__main__':
    app = FullUpdater()
    app.add_command()
    app.run()
